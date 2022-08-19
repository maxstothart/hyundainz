#setup
import os
import sys
import csv
import time
from pdf2image import convert_from_path, convert_from_bytes
import pytesseract
from PIL import Image
from openpyxl import load_workbook
from datetime import date
import pandas as pd
#fin = "input\\CCF_000430"
fin = sys.argv[1] #set argument as variable
wb = load_workbook('Support\\Original.xlsx')#load excel document
coversheet = wb['Cover Sheet']#select sheet 1 from excel document
quote = wb['Quote']# select sheet 2 from excel document
#os.system("python3 dll/pdf2csv.py "+ fin)# run pdf to table (not important yet)
file = open("TEMP/output.csv")#open csv file
csvreader = csv.reader(file)#read csv file
rows = []
for row in csvreader:
        rows.append(row)
file.close()#close csv file

#function definitions
def extract(string, after): #extract string after string function
    global out
    v1 = output.split(after,1)[1]
    n=(v1.find("\n"))
    out = (v1[ 0 : n ])
    time.sleep(0.1)
def celladd(in1, in2):
    global out
    out = ""
    v1=(tables[in1].value)
    v2=(tables[in2].value)
    if not(v2 and v1):
        if(v2):
            out = v2
        elif(v1):
            out = v1
    else:
        out=(v1+v2)

#ocr
pytesseract.pytesseract.tesseract_cmd = r'files/Tesseract-OCR/tesseract'
images = convert_from_path(fin, dpi=200, output_folder="TEMP", fmt="jpg", output_file="output", poppler_path = r"files/poppler-22.04.0/Library/bin")
output = (pytesseract.image_to_string(Image.open('TEMP/output0001-1.jpg')))


#get info from ocr
extract(output, "Parts quote: "); print(out); partsquote=out
extract(output, "Order Type: "); print(out); ordertype=out
extract(output, "Dealer Code: "); print(out); dealercode=out
extract(output, "Partstrader job#: "); print(out); partstraderjob=out
extract(output, "Reference Number: "); print(out); referencenumber=out
extract(output, "Model Code: "); print(out); modelcode=out
extract(output, "VIN: "); print(out); vin=out
extract(output, "Vehicle year: "); print(out); vehicleyear=out
extract(output, "Registration: "); print(out); registration=out
extract(output, "Closing Date: "); print(out); closingdate=out
extract(output, "Status: "); print(out); status=out
extract(output, "Buy total: "); print(out); buytotal=out
extract(output, "Sell total: "); print(out); selltotal=out
extract(output, "Admin comment: "); print(out); admincomment=out
#extract(output, "Address: "); print(out); address=out

#put info into spreadsheet
coversheet['G4'] = date.today()
coversheet['G5'] = partstraderjob
coversheet['G6'] = registration
coversheet['G8'] = modelcode
coversheet['G9'] = vehicleyear
coversheet['G11'] = vin
coversheet['K1'] = admincomment

#table stuff
data= pd.read_csv("TEMP/output.csv")
#print(data)
# Select rows from list using df.loc[df.index[index_list]]
# splitting dataframe by row index
enddata  = data.iloc[12:-6:]
#print(enddata)
enddata.to_excel('TEMP/table.xlsx', index=False)#save table from csv file to table.xlsx
table = load_workbook("TEMP\\table.xlsx", data_only=True)#load table.xlsx
tables = table["Sheet1"]

#combine cells and set as variable
v1=(4+2)
v2=(v1+1)
celladd("a"+str(v1), "a"+str(v2)); out1=(out[1:]); print(out1); partnumber=out1
celladd("b"+str(v1), "b"+str(v2)); print(out); description=out
celladd("c"+str(v1), "c"+str(v2)); print(out); dealer=out
celladd("d"+str(v1), "d"+str(v2)); print(out); hmnz=out
celladd("e"+str(v1), "e"+str(v2)); print(out); oz=out
celladd("f"+str(v1), "f"+str(v2)); print(out); note=out
celladd("g"+str(v1), "g"+str(v2)); print(out); quantity=out
celladd("h"+str(v1), "h"+str(v2)); print(out); buyprice=out
celladd("i"+str(v1), "i"+str(v2)); print(out); sellprice=out

#insert row info into table
v3=10
quote['a'+str(v3)] = description
quote['c'+str(v3)] = partnumber
quote['d'+str(v3)] = quantity
quote['f'+str(v3)] = sellprice
quote['h'+str(v3)] = buyprice



#end
wb.save('output\\clickme.xlsx')#save spreadsheet
